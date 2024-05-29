from pathlib import Path

import openpyxl
import requests
import warnings
import io
from fastapi.responses import FileResponse

warnings.filterwarnings(action='ignore')

excel_list = []
f_KRW_sing = '_-[$₩-ko-KR]* #,##0.00_-;-[$₩-ko-KR]* #,##0.00_-;_-[$₩-ko-KR]* "-"??_-;_-@_-'
f_percent = '0.##"%"'


def exchange_rate_data_get():
    return requests.get("https://open.er-api.com/v6/latest/JPY").json()["rates"]["KRW"]


def excel_exchange(file: io.BytesIO):
    count = 3
    print("Reading Excel file...")
    wb = openpyxl.load_workbook(file)
    sheet = wb.get_sheet_by_name(wb.sheetnames.pop(0))
    for row in sheet.iter_rows(values_only=True):
        excel_list.append(row)
    print("Reading Excel file is done.")
    exchange_rate = exchange_rate_data_get()
    print(f"오늘의 환율은 1JPY당 {exchange_rate}KRW 입니다.")

    for i in excel_list:
        if i == excel_list[0]:
            continue
        elif i == excel_list[1]:
            continue
        name = i[0]
        krw = round(int(i[1]))
        jpy = round(int(i[2]))
        jpy_to_krw = round(jpy * exchange_rate)

        sheet = wb.get_sheet_by_name(wb.sheetnames.pop(1))
        sheet.cell(row=count, column=1).value = name
        sheet.cell(row=count, column=1).style = "main"

        sheet.cell(row=count, column=2).value = krw
        sheet.cell(row=count, column=2).style = "main"
        sheet.cell(row=count, column=2).number_format = f_KRW_sing

        sheet.cell(row=count, column=3).value = jpy_to_krw
        sheet.cell(row=count, column=3).style = "main"
        sheet.cell(row=count, column=3).number_format = f_KRW_sing

        sheet = wb.get_sheet_by_name(wb.sheetnames.pop(2))
        sheet.cell(row=count, column=1).value = name
        sheet.cell(row=count, column=1).style = "main"

        if i[1] - jpy_to_krw < 0:
            sheet.cell(row=count, column=2).value = (krw - jpy_to_krw) * -1
            sheet.cell(row=count, column=2).style = "main"
            sheet.cell(row=count, column=2).number_format = f_KRW_sing
        else:
            sheet.cell(row=count, column=2).value = krw - jpy_to_krw
            sheet.cell(row=count, column=2).style = "main"
            sheet.cell(row=count, column=2).number_format = f_KRW_sing

        if (i[1] - jpy_to_krw) / i[1] * 100 < 0:
            sheet.cell(row=count, column=3).value = round((i[1] - jpy_to_krw) / i[1] * 100 * -1, 2)
            sheet.cell(row=count, column=3).style = "percent"
            sheet.cell(row=count, column=3).number_format = f_percent
        else:
            sheet.cell(row=count, column=3).value = round((i[1] - jpy_to_krw) / i[1] * 100, 2)
            sheet.cell(row=count, column=3).style = "percent"
            sheet.cell(row=count, column=3).number_format = f_percent

        count += 1

    data = io.BytesIO()
    wb.save(data)
    data.seek(0)
    return data


def excel_snack(file: Path):
    print(file)
