from pathlib import Path
from openpyxl.chart import *

import openpyxl
import typer
import requests
import warnings

from typing_extensions import Annotated

warnings.filterwarnings(action='ignore')
app = typer.Typer(name="C$C", help="Currency$Converter")

excel_list = []
f_KRW_sing = '_-[$₩-ko-KR]* #,##0.00_-;-[$₩-ko-KR]* #,##0.00_-;_-[$₩-ko-KR]* "-"??_-;_-@_-'
f_percent = '0.##"%"'


def exchange_rate_data_get():
    return requests.get("https://open.er-api.com/v6/latest/JPY").json()["rates"]["KRW"]


def excel_exchange(file: Path):
    count = 3
    print("Reading Excel file...")
    wb = openpyxl.load_workbook(file)
    sheet = wb.get_sheet_by_name(wb.sheetnames.pop(0))
    for row in sheet.iter_rows(values_only=True):
        excel_list.append(row)
    print("Reading Excel file is done.")
    exchange_rate = exchange_rate_data_get()
    print(f"오늘의 환율은 1JPY당 {exchange_rate}KRW 입니다.")

    for i in excel_list:
        if i == excel_list[0]:
            continue
        elif i == excel_list[1]:
            continue
        name = i[0]
        krw = round(i[1])
        jpy = round(i[2])
        jpy_to_krw = round(jpy * exchange_rate)

        sheet = wb.get_sheet_by_name(wb.sheetnames.pop(1))
        sheet.cell(row=count, column=1).value = name
        sheet.cell(row=count, column=1).style = "main"

        sheet.cell(row=count, column=2).value = krw
        sheet.cell(row=count, column=2).style = "main"
        sheet.cell(row=count, column=2).number_format = f_KRW_sing

        sheet.cell(row=count, column=3).value = jpy_to_krw
        sheet.cell(row=count, column=3).style = "main"
        sheet.cell(row=count, column=3).number_format = f_KRW_sing

        sheet = wb.get_sheet_by_name(wb.sheetnames.pop(2))
        sheet.cell(row=count, column=1).value = name
        sheet.cell(row=count, column=1).style = "main"

        sheet.cell(row=count, column=2).value = krw - jpy_to_krw
        sheet.cell(row=count, column=2).style = "main"
        sheet.cell(row=count, column=2).number_format = f_KRW_sing

        sheet.cell(row=count, column=3).value = round((i[1] - jpy_to_krw) / i[1] * 100, 2)
        sheet.cell(row=count, column=3).style = "percent"
        sheet.cell(row=count, column=3).number_format = f_percent

        count += 1

    chart = BarChart3D()
    chart.type = "col"
    chart.style = 10
    data = Reference(sheet, min_col=3, min_row=2, max_col=3, max_row=len(excel_list))
    categories = Reference(sheet, min_col=1, min_row=3, max_row=len(excel_list))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.title = "한국과 일본의 물가 차이"
    chart.x_axis.title = "품목"
    chart.y_axis.title = "물가 차이(%)"
    chart.legend = None
    chart.y_axis.number_format = f_percent
    sheet.add_chart(chart, "E5")

    wb.save(file)


def excel_snack(file: Path):
    print(file)


@app.command(help="환율 계산기")
def main(file: Annotated[Path, typer.Argument(help="Excel 파일 경로")],
         snack: Annotated[bool, typer.Option(help="과자 계산용")] = False):
    print(snack)
    if snack:
        excel_snack(file)
    else:
        excel_exchange(file)


if __name__ == "__main__":
    app()
